import streamlit as st
import pandas as pd
import spacy
import random
import os
from openpyxl import Workbook, load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score

# Load Datasets
data = pd.read_csv('reduced_emotion_dataset.csv')  # Mood Dataset
affirmations = pd.read_csv('affirmation_new.csv')  # Affirmation Dataset

# Load spaCy model
nlp = spacy.load('en_core_web_sm')

def preprocess_text_spacy(text):
    doc = nlp(text.lower())
    tokens = [token.text for token in doc if not token.is_stop and not token.is_punct]
    return ' '.join(tokens)

data['Cleaned_Text'] = data['Text'].apply(preprocess_text_spacy)

# Feature Extraction
vectorizer = TfidfVectorizer(max_features=5000)
X = vectorizer.fit_transform(data['Cleaned_Text']).toarray()
y = data['Mood']

# Train-Test Split
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Train a Classifier
model = RandomForestClassifier(n_estimators=100, random_state=42)
model.fit(X_train, y_train)

def analyze_mood(user_input):
    cleaned_input = preprocess_text_spacy(user_input)
    input_vector = vectorizer.transform([cleaned_input]).toarray()
    mood_probabilities = model.predict_proba(input_vector)[0]
    mood_classes = model.classes_
    return {mood_classes[i]: mood_probabilities[i] for i in range(len(mood_classes))}

def generate_affirmation(mood_analysis):
    predicted_mood = max(mood_analysis, key=mood_analysis.get)
    filtered_affirmations = affirmations[affirmations['Mood'] == predicted_mood]['Affirmation'].tolist()
    return random.choice(filtered_affirmations) if filtered_affirmations else "Stay positive!"

def save_email_and_affirmation(email, affirmation):
    file_path = 'user_affirmations.xlsx'
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Email", "Affirmation"])
        workbook.save(file_path)
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet.append([email, affirmation])
    workbook.save(file_path)

def main():
    st.title("Mood-Based Affirmation Generator")
    user_input = st.text_area("How are you feeling today? Describe your mood:")
    email = st.text_input("Enter your email (optional):")
    
    if st.button("Get Affirmation"):
        if user_input:
            mood_analysis = analyze_mood(user_input)
            affirmation = generate_affirmation(mood_analysis)
            st.subheader("Mood Analysis")
            for mood, prob in mood_analysis.items():
                st.write(f"{mood}: {prob*100:.2f}%")
            st.subheader("Here's an affirmation for you:")
            st.success(affirmation)
            if email:
                save_email_and_affirmation(email, affirmation)
                st.info("Your affirmation has been saved and will be sent to your email.")
        else:
            st.warning("Please describe your mood to generate an affirmation.")

if __name__ == "__main__":
    main()
